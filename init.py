from  flask import Flask, render_template, request, jsonify,send_file,make_response
import base64
import sqlite3
import os
import threading
import time
import subprocess
import json
import re
from datetime import datetime

from datetime import datetime, timezone
import pytz 


# 导入密码解密模块
from privacy import decrypt_password

app = Flask(__name__)


# 新的公钥参数 (e, n)
public_exponent = 65537
modulus = 11428270940227957444121972623858067884844156065700355794405296166301988372477

# 数据库文件路径
DATABASE = 'server_monitor.db'

def init_db():
    """初始化数据库，创建表"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # 创建主机表，ip、user和port为联合主键
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS hosts (
            ip TEXT NOT NULL,
            user TEXT NOT NULL,
            encrypted_password TEXT NOT NULL,
            port INTEGER DEFAULT 22,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (ip, user, port)
        )
    ''')
    
    # 创建监控日志表 - 用于记录详细的监控指标
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS monitoring_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ip TEXT NOT NULL,
            cpu_usage REAL,
            memory_usage REAL,
            disk_usage REAL,
            network_rx REAL DEFAULT 0,
            network_tx REAL DEFAULT 0,
            is_online INTEGER DEFAULT 0,
            log_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            log_type TEXT DEFAULT 'host_metrics'
        )
    ''')
    
    # 创建索引以加速查询
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_monitoring_logs_time 
        ON monitoring_logs (log_time)
    ''')
    
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_monitoring_logs_ip_time 
        ON monitoring_logs (ip, log_time)
    ''')


    # 创建监控数据表 - 添加网络流量字段
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS monitoring_data (
            ip TEXT NOT NULL,
            cpu_usage REAL,
            memory_usage REAL,
            disk_usage REAL,
            network_rx REAL DEFAULT 0,
            network_tx REAL DEFAULT 0,
            is_online INTEGER DEFAULT 0,
            check_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (ip, check_time)
        )
    ''')
    
    # 创建告警规则表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alert_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_name TEXT NOT NULL,
            host_ip TEXT NOT NULL,
            metric_type TEXT NOT NULL CHECK (metric_type IN ('cpu', 'memory', 'disk')),
            threshold_value REAL NOT NULL,
            comparison_operator TEXT NOT NULL CHECK (comparison_operator IN ('>', '<', '>=', '<=')),
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (host_ip) REFERENCES hosts (ip)
        )
    ''')
    
    # 创建告警通知表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alert_notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_id INTEGER NOT NULL,
            host_ip TEXT NOT NULL,
            metric_type TEXT NOT NULL,
            current_value REAL NOT NULL,
            threshold_value REAL NOT NULL,
            message TEXT NOT NULL,
            severity TEXT NOT NULL CHECK (severity IN ('low', 'medium', 'high', 'critical')),
            is_resolved INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP,
            FOREIGN KEY (rule_id) REFERENCES alert_rules (id)
        )
    ''')
    
    conn.commit()
    conn.close()

@app.route('/')
def index():
    # 获取所有主机信息
    hosts = get_all_hosts()
    return render_template('index.html', hosts=hosts)

@app.route('/add_host', methods=['POST'])
def add_host():
    ip = request.form['ip']
    user = request.form['user']
    password = request.form['pwd']  # 明文密码
    port = request.form.get('port', 22)  # 默认端口22
    
    # 在服务器端进行 RSA 加密
    encrypted_password = rsa_encrypt(password)
    
    # 将主机信息保存到数据库
    success = save_host_to_db(ip, user, encrypted_password, port)
    
    if success:
        return jsonify({'success': True, 'message': '主机添加成功'})
    else:
        return jsonify({'success': False, 'message': '保存失败，可能已存在相同IP、用户名和端口的主机'})

@app.route('/update_host', methods=['POST'])
def update_host():
    old_ip = request.form['old_ip']
    old_user = request.form['old_user']
    old_port = request.form['old_port']
    
    new_ip = request.form['new_ip']
    new_user = request.form['new_user']
    new_port = request.form['new_port']
    new_password = request.form['new_pwd']
    
    # 在服务器端进行 RSA 加密
    encrypted_password = rsa_encrypt(new_password) if new_password else None
    
    # 更新主机信息
    success = update_host_in_db(old_ip, old_user, old_port, new_ip, new_user, new_port, encrypted_password)
    
    if success:
        return jsonify({'success': True, 'message': '主机更新成功'})
    else:
        return jsonify({'success': False, 'message': '更新失败'})

@app.route('/delete_host', methods=['POST'])
def delete_host():
    ip = request.form['ip']
    user = request.form['user']
    port = request.form['port']
    confirm_user = request.form['confirm_user']
    
    # 验证用户名
    if confirm_user != user:
        return jsonify({'success': False, 'message': '用户名不匹配，删除失败'})
    
    # 从数据库删除主机
    success = delete_host_from_db(ip, user, port)
    
    if success:
        return jsonify({'success': True, 'message': '主机删除成功'})
    else:
        return jsonify({'success': False, 'message': '删除失败'})

@app.route('/monitoring_data')
def show_monitoring_data():
    """显示监控数据"""
    data = get_monitoring_data()
    html = "<h1>监控数据</h1><table border='1'><tr><th>IP</th><th>CPU使用率</th><th>内存使用率</th><th>磁盘使用率</th><th>检查时间</th></tr>"
    for row in data:
        html += f"<tr><td>{row[0]}</td><td>{row[1]:.2f}%</td><td>{row[2]:.2f}%</td><td>{row[3]:.2f}%</td><td>{row[4]}</td></tr>"
    html += "</table><br><a href='/'>返回主页</a>"
    return html

def rsa_encrypt(message):
    """使用 RSA 公钥加密消息"""
    try:
        # 将消息转换为整数
        m = text_to_int(message)
        
        # 检查消息是否小于模数
        if m >= modulus:
            raise ValueError("消息太长，无法使用当前密钥加密")
        
        # 计算 c = m^e mod n
        c = pow(m, public_exponent, modulus)
        
        # 将加密结果转换为 Base64 以便存储/传输
        return base64.b64encode(c.to_bytes((c.bit_length() + 7) // 8, 'big')).decode()
    except Exception as e:
        print(f"加密错误: {e}")
        return f"加密失败: {str(e)}"

def text_to_int(text):
    """将文本转换为整数"""
    n = 0
    for char in text:
        n = n * 256 + ord(char)
    return n

def save_host_to_db(ip, user, encrypted_password, port=22):
    """将主机信息保存到数据库"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 使用 INSERT OR REPLACE 来处理重复键（更新密码）
        cursor.execute('''
            INSERT OR REPLACE INTO hosts (ip, user, encrypted_password, port)
            VALUES (?, ?, ?, ?)
        ''', (ip, user, encrypted_password, port))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return False

def update_host_in_db(old_ip, old_user, old_port, new_ip, new_user, new_port, encrypted_password):
    """更新主机信息"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 如果密码为空，则只更新IP、用户名和端口
        if encrypted_password:
            cursor.execute('''
                UPDATE hosts 
                SET ip = ?, user = ?, port = ?, encrypted_password = ?
                WHERE ip = ? AND user = ? AND port = ?
            ''', (new_ip, new_user, new_port, encrypted_password, old_ip, old_user, old_port))
        else:
            cursor.execute('''
                UPDATE hosts 
                SET ip = ?, user = ?, port = ?
                WHERE ip = ? AND user = ? AND port = ?
            ''', (new_ip, new_user, new_port, old_ip, old_user, old_port))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return False

def delete_host_from_db(ip, user, port):
    """从数据库删除主机"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM hosts WHERE ip = ? AND user = ? AND port = ?', (ip, user, port))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return False

def get_all_hosts():
    """获取所有主机信息"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('SELECT ip, user, encrypted_password, port, created_at FROM hosts')
        hosts = cursor.fetchall()
        
        conn.close()
        return hosts
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return []

def get_all_hosts_with_monitoring():
    """获取所有主机及其监控配置数据"""
    try:
        # 获取所有主机基本信息
        hosts = get_all_hosts()
        
        # 获取所有主机的最新监控数据
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 获取每个主机的最新监控数据
        cursor.execute('''
            SELECT md.ip, md.cpu_usage, md.memory_usage, md.disk_usage, 
                   md.network_rx, md.network_tx, md.is_online, md.check_time
            FROM monitoring_data md
            INNER JOIN (
                SELECT ip, MAX(check_time) as latest_time
                FROM monitoring_data
                GROUP BY ip
            ) latest ON md.ip = latest.ip AND md.check_time = latest.latest_time
        ''')
        
        monitoring_data = cursor.fetchall()
        
        # 检查是否存在主机监控配置表
        cursor.execute('''
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='host_monitor_configs'
        ''')
        config_table_exists = cursor.fetchone() is not None
        
        # 获取主机监控配置（如果表存在）
        host_configs = {}
        if config_table_exists:
            cursor.execute('''
                SELECT ip, user, port, monitor_cpu, monitor_memory, 
                       monitor_disk, monitor_network, status
                FROM host_monitor_configs
            ''')
            configs = cursor.fetchall()
            for config in configs:
                ip, user, port, cpu, memory, disk, network, status = config
                key = f"{ip}-{user}-{port}"
                host_configs[key] = {
                    'monitor_cpu': bool(cpu),
                    'monitor_memory': bool(memory),
                    'monitor_disk': bool(disk),
                    'monitor_network': bool(network),
                    'status': status
                }
        
        conn.close()
        
        # 将监控数据转换为字典，便于查找
        monitoring_dict = {}
        for data in monitoring_data:
            ip, cpu, memory, disk, network_rx, network_tx, is_online, check_time = data
            monitoring_dict[ip] = {
                'cpu_usage': cpu if cpu is not None else 0,
                'memory_usage': memory if memory is not None else 0,
                'disk_usage': disk if disk is not None else 0,
                'network_rx': network_rx if network_rx is not None else 0,
                'network_tx': network_tx if network_tx is not None else 0,
                'is_online': bool(is_online) if is_online is not None else False,
                'check_time': check_time
            }
        
        # 合并主机信息和监控数据
        result = {}
        for host in hosts:
            ip = host[0]
            user = host[1]
            port = host[3]
            
            host_key = f"{ip}-{user}-{port}"
            
            # 获取该主机的监控数据（如果存在）
            host_monitoring = monitoring_dict.get(ip, {
                'cpu_usage': 0,
                'memory_usage': 0,
                'disk_usage': 0,
                'network_rx': 0,
                'network_tx': 0,
                'is_online': False,
                'check_time': None
            })
            
            # 获取该主机的监控配置（如果存在）
            host_config = host_configs.get(host_key, {
                'monitor_cpu': True,
                'monitor_memory': True,
                'monitor_disk': True,
                'monitor_network': True,
                'status': 'active'
            })
            
            # 合并数据
            result[host_key] = {
                'ip': ip,
                'user': user,
                'port': port,
                'status': host_config['status'],
                'monitor_cpu': host_config['monitor_cpu'],
                'monitor_memory': host_config['monitor_memory'],
                'monitor_disk': host_config['monitor_disk'],
                'monitor_network': host_config['monitor_network'],
                'is_online': host_monitoring['is_online'],
                'cpu_usage': host_monitoring['cpu_usage'],
                'memory_usage': host_monitoring['memory_usage'],
                'disk_usage': host_monitoring['disk_usage'],
                'network_rx': host_monitoring['network_rx'],
                'network_tx': host_monitoring['network_tx'],
                'last_check': host_monitoring['check_time']
            }
        
        return result
        
    except Exception as e:
        print(f"获取主机监控数据失败: {e}")
        return {}


def get_monitoring_data():
    """获取监控数据"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('SELECT ip, cpu_usage, memory_usage, disk_usage, check_time FROM monitoring_data ORDER BY check_time DESC LIMIT 100')
        data = cursor.fetchall()
        
        conn.close()
        return data
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return []

def save_monitoring_data(ip, cpu_usage, memory_usage, disk_usage, is_online=0, network_rx=0, network_tx=0):
    """保存监控数据到数据库"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO monitoring_data (ip, cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (ip, cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return False

def check_alerts(ip, cpu_usage, memory_usage, disk_usage):
    """检查告警条件并创建通知"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 获取该主机的所有活跃告警规则
        cursor.execute('''
            SELECT id, rule_name, metric_type, threshold_value, comparison_operator
            FROM alert_rules
            WHERE host_ip = ? AND is_active = 1
        ''', (ip,))
        
        rules = cursor.fetchall()
        
        # 检查每个告警规则
        for rule in rules:
            rule_id, rule_name, metric_type, threshold_value, comparison_operator = rule
            
            # 确定当前值和严重程度
            current_value = None
            if metric_type == 'cpu':
                current_value = cpu_usage
            elif metric_type == 'memory':
                current_value = memory_usage
            elif metric_type == 'disk':
                current_value = disk_usage
            
            if current_value is None:
                continue
            
            # 检查告警条件
            is_triggered = False
            if comparison_operator == '>':
                is_triggered = current_value > threshold_value
            elif comparison_operator == '<':
                is_triggered = current_value < threshold_value
            elif comparison_operator == '>=':
                is_triggered = current_value >= threshold_value
            elif comparison_operator == '<=':
                is_triggered = current_value <= threshold_value
            
            if is_triggered:
                # 检查是否已存在未解决的相同告警
                cursor.execute('''
                    SELECT id FROM alert_notifications
                    WHERE rule_id = ? AND host_ip = ? AND is_resolved = 0
                    ORDER BY created_at DESC LIMIT 1
                ''', (rule_id, ip))
                
                existing_alert = cursor.fetchone()
                
                # 如果没有未解决的告警，则创建新的告警通知
                if not existing_alert:
                    # 确定严重程度
                    severity = 'medium'
                    if current_value >= 90:
                        severity = 'critical'
                    elif current_value >= 80:
                        severity = 'high'
                    elif current_value >= 70:
                        severity = 'medium'
                    else:
                        severity = 'low'
                    
                    # 创建告警消息
                    metric_names = {'cpu': 'CPU使用率', 'memory': '内存使用率', 'disk': '磁盘使用率'}
                    message = f"主机 {ip} {metric_names[metric_type]} 超过阈值！当前值: {current_value:.2f}%, 阈值: {threshold_value:.2f}%"
                    
                    # 插入告警通知
                    cursor.execute('''
                        INSERT INTO alert_notifications (rule_id, host_ip, metric_type, current_value, threshold_value, message, severity)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (rule_id, ip, metric_type, current_value, threshold_value, message, severity))
                    
                    print(f"告警触发: {message}")
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"告警检查错误: {e}")
        return False

def create_alert_rule(rule_name, host_ip, metric_type, threshold_value, comparison_operator):
    """创建告警规则"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO alert_rules (rule_name, host_ip, metric_type, threshold_value, comparison_operator)
            VALUES (?, ?, ?, ?, ?)
        ''', (rule_name, host_ip, metric_type, threshold_value, comparison_operator))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"创建告警规则错误: {e}")
        return False

def get_alert_rules():
    """获取所有告警规则"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT ar.id, ar.rule_name, ar.host_ip, ar.metric_type, ar.threshold_value, 
                   ar.comparison_operator, ar.is_active, ar.created_at
            FROM alert_rules ar
            ORDER BY ar.created_at DESC
        ''')
        
        rules = cursor.fetchall()
        conn.close()
        return rules
    except sqlite3.Error as e:
        print(f"获取告警规则错误: {e}")
        return []

def get_active_alerts():
    """获取未解决的活跃告警"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT an.id, an.rule_id, an.host_ip, an.metric_type, an.current_value,
                   an.threshold_value, an.message, an.severity, an.created_at
            FROM alert_notifications an
            WHERE an.is_resolved = 0
            ORDER BY an.created_at DESC
        ''')
        
        alerts = cursor.fetchall()
        conn.close()
        return alerts
    except sqlite3.Error as e:
        print(f"获取活跃告警错误: {e}")
        return []

def resolve_alert(alert_id):
    """解决告警"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE alert_notifications
            SET is_resolved = 1, resolved_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (alert_id,))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"解决告警错误: {e}")
        return False

def delete_alert_rule(rule_id):
    """删除告警规则"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM alert_rules WHERE id = ?', (rule_id,))
        cursor.execute('UPDATE alert_notifications SET is_resolved = 1 WHERE rule_id = ?', (rule_id,))
        
        conn.commit()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"删除告警规则错误: {e}")
        return False

def collect_monitoring_data():
    """采集监控数据"""
    # 导入配置变量 - 使用默认值（True）来启用所有监控项
    monitor_cpu = True
    monitor_memory = True
    monitor_disk = True
    monitor_network = True
    hosts = get_all_hosts()
    
    for host in hosts:
        ip, user, encrypted_password, port, _ = host
        
        # 解密密码
        password = decrypt_password(encrypted_password)
        if not password:
            print(f"无法解密密码 for {ip}, 跳过")
            continue
        
        # 使用Ansible采集数据
        try:
            # 创建临时inventory文件
            inventory_content = f"""
[all]
{ip}

[all:vars]
ansible_connection=ssh
ansible_ssh_user={user}
ansible_ssh_pass={password}
ansible_ssh_port={port}
ansible_become_pass={password}
"""
            
            with open('ansible_inventory', 'w') as f:
                f.write(inventory_content)
            
            # 执行ping命令检查上线情况
            ping_cmd = ["ansible", ip, "-i", "ansible_inventory", "-m", "ping"]
            result = subprocess.run(ping_cmd, capture_output=True, text=True)
            output = result.stdout.strip()
            
            if "SUCCESS" in output and "pong" in output:
                print(f"{ip} | 成功响应")
                is_online = 1
                
                # 根据配置采集CPU使用率
                if monitor_cpu:
                    try:
                        # 执行Ansible命令获取CPU使用率
                        cpu_cmd = ["ansible", ip, "-i", "ansible_inventory", "-m", "shell", 
                                    "-a", "grep 'cpu ' /proc/stat | awk '{usage=($2+$4)*100/($2+$4+$5)} END {print usage}'"]
                        cpu_result = subprocess.run(cpu_cmd, capture_output=True, text=True)
                        cpu_output = cpu_result.stdout.strip().split('\n')[-1]
                        
                        # 处理Ansible输出格式
                        if "SUCCESS" in cpu_result.stdout and "rc=0" in cpu_result.stdout:
                            # 提取实际的CPU使用率值
                            lines = cpu_result.stdout.split('\n')
                            for line in lines:
                                if line.strip().replace('.', '').replace('%', '').isdigit():
                                    cpu_usage = float(line.strip())
                                    break
                            else:
                                cpu_usage = 0.0
                        else:
                            # 如果Ansible命令执行成功但输出格式不同
                            if cpu_result.returncode == 0 and cpu_output.replace('.', '').isdigit():
                                cpu_usage = float(cpu_output)
                            else:
                                print(f"无法解析 {ip} 的CPU使用率: {cpu_output}")
                                cpu_usage = 0.0
                    except Exception as e:
                        print(f"采集 {ip} CPU使用率时出错: {e}")
                        cpu_usage = 0.0
                else:
                    cpu_usage = 0.0  # 如果未配置监控CPU，则设为0

                
                # 根据配置采集内存使用率
                if monitor_memory:
                    try:
                        # 执行Ansible命令获取内存使用率
                        memory_cmd = ["ansible", ip, "-i", "ansible_inventory", "-m", "shell",  
                                      "-a", "free | grep Mem | awk '{print $3/$2 * 100.0}'"]
                        memory_result = subprocess.run(memory_cmd, capture_output=True, text=True)
                        memory_output = memory_result.stdout.strip().split('\n')[-1]
                        memory_usage = float(memory_output) if memory_result.returncode == 0 and memory_output.replace('.', '').isdigit() else 0.0
                    except Exception as e:
                        print(f"采集 {ip} 内存使用率时出错: {e}")
                        memory_usage = 0.0
                else:
                    memory_usage = 0.0  # 如果未配置监控内存，则设为0

                
                                # 根据配置采集磁盘使用率
                if monitor_disk:
                    try:
                        # 执行Ansible命令获取磁盘使用率
                        disk_cmd = ["ansible", ip, "-i", "ansible_inventory", "-m", "shell", 
                                    "-a", "df / | awk 'NR==2 {print $5}' | sed 's/%//'"]
                        disk_result = subprocess.run(disk_cmd, capture_output=True, text=True)
                        disk_output = disk_result.stdout.strip().split('\n')[-1]
                        disk_usage = float(disk_output) if disk_result.returncode == 0 and disk_output.replace('.', '').isdigit() else 0.0
                    except Exception as e:
                        print(f"采集 {ip} 磁盘使用率时出错: {e}")
                        disk_usage = 0.0
                else:
                    disk_usage = 0.0  # 如果未配置监控磁盘，则设为0

                
                # ========== 网络流量数据采集 ==========
                if monitor_network:
                    try:
                        print(f"正在采集 {ip} 的网络流量数据...")
                        
                        # 使用更精确的awk命令直接提取接收和发送字节数
                        network_cmd = ["ansible", ip, "-i", "ansible_inventory", "-m", "shell", 
                                    "-a", "cat /proc/net/dev | awk '/ens33:/ {print $2, $10}'"]
                        
                        network_result = subprocess.run(network_cmd, capture_output=True, text=True)
                        network_output = network_result.stdout.strip()
                        
                        print(f"网络流量命令输出: {repr(network_output)}")
                        
                        if network_result.returncode == 0 and network_output:
                            # 解析输出，提取数字
                            lines = network_output.split('\n')
                            for line in lines:
                                if line.strip() and not line.startswith('192.168'):  # 跳过IP地址行
                                    numbers = re.findall(r'\d+', line)
                                    if len(numbers) >= 2:
                                        try:
                                            rx_bytes = int(numbers[0])
                                            tx_bytes = int(numbers[1])
                                            network_rx = rx_bytes / 1024 / 1024  # 转换为MB
                                            network_tx = tx_bytes / 1024 / 1024  # 转换为MB
                                            print(f"成功采集 {ip} 的网络流量: RX={network_rx:.2f}MB, TX={network_tx:.2f}MB")
                                            break
                                        except ValueError as e:
                                            print(f"解析网络流量数据失败: {e}")
                                            continue
                            else:
                                # 如果解析失败，使用备用方法：直接查看/proc/net/dev完整输出
                                print("尝试备用方法...")
                                debug_cmd = ["ansible", ip, "-i", "ansible_inventory", "-m", "shell", 
                                            "-a", "cat /proc/net/dev | grep ens33:"]
                                debug_result = subprocess.run(debug_cmd, capture_output=True, text=True)
                                debug_output = debug_result.stdout.strip()
                                print(f"调试输出: {repr(debug_output)}")
                                
                                # 从调试输出中手动提取
                                if debug_result.returncode == 0 and debug_output:
                                    # 提取所有数字
                                    all_numbers = re.findall(r'\d+', debug_output)
                                    if len(all_numbers) >= 2:
                                        # 假设前两个数字是接收和发送字节数
                                        rx_bytes = int(all_numbers[0])
                                        tx_bytes = int(all_numbers[1])
                                        network_rx = rx_bytes / 1024 / 1024
                                        network_tx = tx_bytes / 1024 / 1024
                                        print(f"备用方法成功: RX={network_rx:.2f}MB, TX={network_tx:.2f}MB")
                                    else:
                                        print("备用方法失败，使用模拟数据")
                                        network_rx = 0.5
                                        network_tx = 0.3
                                else:
                                    print("调试命令失败，使用模拟数据")
                                    network_rx = 0.5
                                    network_tx = 0.3
                        else:
                            print(f"网络流量命令执行失败，返回码: {network_result.returncode}")
                            network_rx = 0.5
                            network_tx = 0.3
                            
                    except Exception as e:
                        print(f"采集 {ip} 网络流量数据时出错: {e}")
                        network_rx = 0.5
                        network_tx = 0.3
                else:
                    network_rx = 0.0
                    network_tx = 0.0

                # ========== 网络流量数据采集结束 ==========
                
                print(f"采集到 {ip} 的数据: CPU={cpu_usage:.2f}%, 内存={memory_usage:.2f}%, 磁盘={disk_usage:.2f}%")
                
                # 保存到数据库 - 添加网络流量参数
                save_monitoring_data(ip, cpu_usage, memory_usage, disk_usage, is_online, network_rx, network_tx)
                
                # 检查告警
                check_alerts(ip, cpu_usage, memory_usage, disk_usage)
                
            else:
                print(f"{ip} | 未成功响应: {output}")
                # 保存离线状态，其他指标设为0
                is_online = 0
                save_monitoring_data(ip, 0.0, 0.0, 0.0, is_online, 0.0, 0.0)
                
        except Exception as e:
            print(f"采集 {ip} 数据时出错: {e}")
            # 保存错误状态
            save_monitoring_data(ip, 0.0, 0.0, 0.0, 0, 0.0, 0.0)
    
        # 删除临时文件
    try:
        os.remove('ansible_inventory')
    except:
        pass
    
    # 写入主机监控指标到日志文件
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 使用美化的日志格式
        log_header = f"\n{'='*80}\n{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - 监控数据采集完成\n{'='*80}\n"
        with open(os.environ.get('MONITOR_LOG_FILE', 'monitor.log'), 'a', encoding='utf-8') as f:
            f.write(log_header)
            
            # 获取所有主机的最新监控数据
            hosts = get_all_hosts()
            
            # 为每个主机写入详细的监控指标
            for host in hosts:
                ip = host[0]
                cursor.execute('''
                    SELECT cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online, check_time
                    FROM monitoring_data 
                    WHERE ip = ? 
                    ORDER BY check_time DESC 
                    LIMIT 1
                ''', (ip,))
                
                result = cursor.fetchone()
                if result:
                    cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online, check_time = result
                    status = "✅ 在线" if is_online else "❌ 离线"
                    
                    # 美化的主机日志格式
                    host_log = f"""
┌─────────────────────────────────────────────────────────────────┐
│ 主机: {ip:<55} │
│ 状态: {status:<55} │
│ CPU:    {cpu_usage:>6.2f}%  内存:  {memory_usage:>6.2f}%  磁盘: {disk_usage:>6.2f}% │
│ 网络RX: {network_rx:>6.2f}MB  网络TX: {network_tx:>6.2f}MB            │
└─────────────────────────────────────────────────────────────────┘
"""
                    f.write(host_log)
            
            f.write("\n")  # 添加空行分隔不同时间点的采集
            
        conn.close()
    except Exception as e:
        print(f"写入监控日志时出错: {e}")





def cleanup_old_data(days=30):
    """清理指定天数前的监控数据"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            DELETE FROM monitoring_data 
            WHERE check_time < date('now', '-{} days')
        '''.format(days))
        
        deleted_rows = cursor.rowcount
        conn.commit()
        conn.close()
        
        print(f"已清理{deleted_rows}条{days}天前的监控数据")
        return True
    except sqlite3.Error as e:
        print(f"清理旧数据失败: {e}")
        return False




def cleanup_old_data(days=30):
    """清理指定天数前的监控数据"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute('''
            DELETE FROM monitoring_data 
            WHERE check_time < date('now', '-{} days')
        '''.format(days))
        
        deleted_rows = cursor.rowcount
        conn.commit()
        conn.close()
        
        print(f"已清理{deleted_rows}条{days}天前的监控数据")
        return True
    except sqlite3.Error as e:
        print(f"清理旧数据失败: {e}")
        return False

def monitoring_loop():
    """监控数据采集循环"""
    last_cleanup_time = time.time() - (24 * 60 * 60)
    
    while True:
        try:
            if globals().get('stop_monitoring', False):
                break
                
            # 直接读取环境变量，确保获取最新配置
            monitor_interval = int(os.environ.get('MONITOR_INTERVAL', 30))
            cleanup_days = int(os.environ.get('MONITOR_CLEANUP_DAYS', 30))
            enable_auto_cleanup = os.environ.get('MONITOR_AUTO_CLEANUP', 'True').lower() == 'true'
            
            print(f"{datetime.now()}: 开始采集监控数据")
            
            # 采集监控数据
            collect_monitoring_data()
            
            current_time = time.time()
            print(f"{datetime.now()}: 监控数据采集完成，等待{monitor_interval}秒")
            
            # 定期清理旧数据
            if enable_auto_cleanup and (current_time - last_cleanup_time) >= 24 * 60 * 60:
                cleanup_old_data(cleanup_days)
                last_cleanup_time = current_time
                print(f"{datetime.now()}: 已清理{cleanup_days}天前的旧数据")
            
            time.sleep(monitor_interval)
            
        except Exception as e:
            print(f"监控循环出错: {e}")
            time.sleep(30)


# 添加一个路由来查看所有主机（用于调试）
@app.route('/hosts')
def show_hosts():
    hosts = get_all_hosts()
    html = "<h1>已添加的主机</h1><table border='1'><tr><th>IP</th><th>用户名</th><th>端口</th><th>加密密码</th><th>添加时间</th></tr>"
    for host in hosts:
        html += f"<tr><td>{host[0]}</td><td>{host[1]}</td><td>{host[3]}</td><td>{host[2][:20]}...</td><td>{host[4]}</td></tr>"
    html += "</table><br><a href='/'>返回主页</a>"
    return html


@app.route('/api/monitor_management_data')
def api_monitor_management_data():
    """统一监控管理页面数据API"""
    try:
        # 获取所有数据
        hosts_data = get_all_hosts_with_monitoring()
        collection_status = get_collection_status_data()  # 使用新的函数
        recent_logs_data = get_logs(limit=50)  # 这个已经返回字典
        
        # 构建统一响应
        result = {
            'success': True,
            'hosts': hosts_data,
            'collection_status': collection_status,
            'recent_logs': recent_logs_data if recent_logs_data.get('success') else {
                'logs': ['日志数据获取失败'],
                'count': 1,
                'type': 'error'
            },
            'timestamp': datetime.now().isoformat()
        }
        
        return jsonify(result)
        
    except Exception as e:
        print(f"统一API错误: {e}")
        return jsonify({
            'success': False, 
            'message': str(e),
            'hosts': {},
            'collection_status': {
                'status': '错误',
                'is_active': False,
                'last_collection_time': '未知',
                'collection_interval': '未知',
                'next_collection_time': '未知'
            },
            'recent_logs': {
                'logs': [f'API错误: {str(e)}'],
                'count': 1,
                'type': 'error'
            }
        })


@app.route('/get_monitoring_logs')
def get_monitoring_logs():
    """获取监控日志，支持按时间范围和主机过滤"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date', '')  # 格式: YYYY-MM-DD
        end_date = request.args.get('end_date', '')      # 格式: YYYY-MM-DD
        ip_filter = request.args.get('ip', '')           # 主机IP过滤
        page = int(request.args.get('page', 1))           # 页码
        per_page = int(request.args.get('per_page', 50)) # 每页条数
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 构建查询条件
        conditions = []
        params = []
        
        if start_date:
            conditions.append("DATE(log_time) >= ?")
            params.append(start_date)
            
        if end_date:
            conditions.append("DATE(log_time) <= ?")
            params.append(end_date)
            
        if ip_filter:
            conditions.append("ip = ?")
            params.append(ip_filter)
            
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        
        # 获取总记录数
        count_query = f"SELECT COUNT(*) FROM monitoring_logs {where_clause}"
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]
        
        # 计算分页
        offset = (page - 1) * per_page
        total_pages = (total_records + per_page - 1) // per_page
        
        # 查询数据
        query = f"""
            SELECT ip, cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online, log_time
            FROM monitoring_logs 
            {where_clause}
            ORDER BY log_time DESC
            LIMIT ? OFFSET ?
        """
        
        cursor.execute(query, params + [per_page, offset])
        logs = cursor.fetchall()
        
        conn.close()
        
        # 格式化数据
        formatted_logs = []
        for log in logs:
            formatted_logs.append({
                'ip': log[0],
                'cpu_usage': log[1],
                'memory_usage': log[2],
                'disk_usage': log[3],
                'network_rx': log[4],
                'network_tx': log[5],
                'is_online': bool(log[6]),
                'log_time': log[7]
            })
        
        # 获取所有主机IP用于过滤下拉框
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT ip FROM hosts ORDER BY ip")
        hosts = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({
            'success': True,
            'logs': formatted_logs,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_records': total_records,
                'total_pages': total_pages
            },
            'hosts': hosts
        })
        
    except Exception as e:
        print(f"获取监控日志失败: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/monitor_dashboard')
def monitor_dashboard():
    """监控大屏页面"""
    return render_template('monitor_dashboard.html')

@app.route('/monitor_management')
def monitor_management():
    """监控管理页面"""
    return render_template('monitor_management.html')

@app.route('/get_collection_status')
def get_collection_status():
    """获取采集状态信息（路由版本）"""
    data = get_collection_status_data()
    return jsonify(data)
@app.route('/api/hosts')
def api_hosts():
    """获取所有主机数据的API端点，供监控管理页面使用"""
    try:
        # 获取所有主机基本信息
        hosts = get_all_hosts()
        
        # 获取所有主机的最新监控数据
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 获取每个主机的最新监控数据
        cursor.execute('''
            SELECT md.ip, md.cpu_usage, md.memory_usage, md.disk_usage, 
                   md.network_rx, md.network_tx, md.is_online, md.check_time
            FROM monitoring_data md
            INNER JOIN (
                SELECT ip, MAX(check_time) as latest_time
                FROM monitoring_data
                GROUP BY ip
            ) latest ON md.ip = latest.ip AND md.check_time = latest.latest_time
        ''')
        
        monitoring_data = cursor.fetchall()
        conn.close()
        
        # 将监控数据转换为字典，便于查找
        monitoring_dict = {}
        for data in monitoring_data:
            ip, cpu, memory, disk, network_rx, network_tx, is_online, check_time = data
            monitoring_dict[ip] = {
                'cpu_usage': cpu,
                'memory_usage': memory,
                'disk_usage': disk,
                'network_rx': network_rx,
                'network_tx': network_tx,
                'is_online': bool(is_online),
                'check_time': check_time
            }
        
        # 合并主机信息和监控数据
        result = {}
        for host in hosts:
            # host 是一个元组：(ip, user, encrypted_password, port, created_at)
            ip = host[0]
            user = host[1]
            port = host[3]
            created_at = host[4]
            
            host_key = f"{ip}-{user}-{port}"
            
            # 获取该主机的监控数据（如果存在）
            host_monitoring = monitoring_dict.get(ip, {
                'cpu_usage': 0,
                'memory_usage': 0,
                'disk_usage': 0,
                'network_rx': 0,
                'network_tx': 0,
                'is_online': False,
                'check_time': None
            })
            
            # 合并数据
            result[host_key] = {
                'ip': ip,
                'user': user,
                'port': port,
                'created_at': created_at,
                'status': 'active',  # 默认启用
                'monitor_cpu': True,
                'monitor_memory': True,
                'monitor_disk': True,
                'monitor_network': True,
                # 添加监控数据
                'is_online': host_monitoring['is_online'],
                'cpu_usage': host_monitoring['cpu_usage'],
                'memory_usage': host_monitoring['memory_usage'],
                'disk_usage': host_monitoring['disk_usage'],
                'network_rx': host_monitoring['network_rx'],
                'network_tx': host_monitoring['network_tx'],
                'last_check': host_monitoring['check_time']
            }
        
        return jsonify({'success': True, 'hosts': result})
    except Exception as e:
        print(f"获取主机数据失败: {e}")
        return jsonify({'success': False, 'message': str(e), 'hosts': {}})

def get_logs(limit=50):
    """获取最近的日志，支持limit参数"""
    try:
        log_file = os.environ.get('MONITOR_LOG_FILE', 'monitor.log')
        lines = []
        
        # 确保日志文件存在
        if not os.path.exists(log_file):
            # 创建日志文件并写入初始内容
            with open(log_file, 'w') as f:
                f.write(f"{datetime.now()}: 监控日志初始化\n")
        
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            # 修改这里：使用传入的limit参数，而不是固定50行
            lines = lines[-limit:]
        
        # 如果日志为空，返回默认消息
        if not lines:
            lines = ["暂无日志记录\n"]
        
        # 解析主机监控数据
        host_logs = []
        system_logs = []
        
        for line in lines:
            # 检查是否是主机监控数据
            if "主机" in line and "状态" in line and "CPU" in line:
                host_logs.append(line.strip())
            else:
                system_logs.append(line.strip())
        
        # 优先返回主机监控数据
        if host_logs:
            return {
                'success': True,
                'logs': host_logs,
                'count': len(host_logs),
                'type': 'host_metrics'
            }
        else:
            return {
                'success': True,
                'logs': system_logs,
                'count': len(system_logs),
                'type': 'system'
            }
    except Exception as e:
        return {
            'success': False,
            'message': str(e),
            'logs': [f"获取日志失败: {str(e)}\n"],
            'count': 1,
            'type': 'error'
        }


def get_collection_status_data():
    """获取采集状态信息（返回字典数据）"""
    try:
        # 检查监控线程是否运行
        is_active = False
        if 'monitor_thread' in globals():
            is_active = monitor_thread.is_alive()
        
        # 获取最后采集时间
        last_collection_time = "无记录"
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('SELECT MAX(check_time) FROM monitoring_data')
        result = cursor.fetchone()
        if result and result[0]:
            last_collection_time = result[0]
        conn.close()
        
        # 获取配置信息
        interval = os.environ.get('MONITOR_INTERVAL', '300')
        
        # 获取默认监控项配置
        monitor_items = []
        if os.environ.get('MONITOR_CPU', 'true').lower() == 'true':
            monitor_items.append('CPU')
        if os.environ.get('MONITOR_MEMORY', 'true').lower() == 'true':
            monitor_items.append('内存')
        if os.environ.get('MONITOR_DISK', 'true').lower() == 'true':
            monitor_items.append('磁盘')
        if os.environ.get('MONITOR_NETWORK', 'true').lower() == 'true':
            monitor_items.append('网络')
        
        # 计算下次采集时间
        next_collection_time = "未知"
        if last_collection_time != "无记录" and is_active:
            try:
                import datetime
                last_time = datetime.datetime.strptime(last_collection_time, '%Y-%m-%d %H:%M:%S')
                next_time = last_time + datetime.timedelta(seconds=int(interval))
                next_collection_time = next_time.strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass
        
        return {
            'status': '活跃' if is_active else '未运行',
            'is_active': is_active,
            'last_collection_time': last_collection_time,
            'collection_interval': f"{interval}秒",
            'next_collection_time': next_collection_time,
            'monitor_items': monitor_items
        }
    except Exception as e:
        return {
            'status': '错误',
            'is_active': False,
            'last_collection_time': '无记录',
            'collection_interval': '未知',
            'next_collection_time': '未知',
            'error': str(e),
            'monitor_items': []
        }


@app.route('/get_recent_logs')
def get_recent_logs():
    """获取最近的监控日志，支持limit参数"""
    try:
        limit = request.args.get('limit', 50, type=int)
        logs_data = get_logs(limit=limit)  # 获取字典数据
        return jsonify(logs_data)  # 转换为Response对象
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': str(e),
            'logs': ['获取日志失败'],
            'count': 1,
            'type': 'error'
        })


@app.route('/get_real_time_data')
def get_real_time_data():
    """获取所有主机的实时监控数据"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 获取每个主机的最新监控数据
        cursor.execute('''
            SELECT md.ip, md.cpu_usage, md.memory_usage, md.disk_usage, 
                   md.network_rx, md.network_tx, md.is_online, md.check_time
            FROM monitoring_data md
            INNER JOIN (
                SELECT ip, MAX(check_time) as latest_time
                FROM monitoring_data
                GROUP BY ip
            ) latest ON md.ip = latest.ip AND md.check_time = latest.latest_time
        ''')
        
        data = cursor.fetchall()
        conn.close()
        
        # 转换为字典格式便于前端使用
        result = {}
        for row in data:
            ip, cpu, memory, disk, network_rx, network_tx, is_online, check_time = row
            result[ip] = {
                'cpu_usage': cpu if cpu is not None else 0,
                'memory_usage': memory if memory is not None else 0,
                'disk_usage': disk if disk is not None else 0,
                'network_rx': network_rx if network_rx is not None else 0,
                'network_tx': network_tx if network_tx is not None else 0,
                'is_online': bool(is_online) if is_online is not None else False,
                'check_time': check_time
            }
        
        return jsonify(result)
    except sqlite3.Error as e:
        print(f"数据库错误: {e}")
        return jsonify({})  # 返回空对象而不是错误




    """获取最近的日志，优先返回主机监控指标"""
    try:
        log_file = os.environ.get('MONITOR_LOG_FILE', 'monitor.log')
        lines = []
        
        # 确保日志文件存在
        if not os.path.exists(log_file):
            # 创建日志文件并写入初始内容
            with open(log_file, 'w') as f:
                f.write(f"{datetime.now()}: 监控日志初始化\n")
        
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            # 返回最后50行（增加行数以显示更多主机监控数据）
            lines = lines[-50:]
        
        # 如果日志为空，返回默认消息
        if not lines:
            lines = ["暂无日志记录\n"]
        
        # 解析主机监控数据
        host_logs = []
        system_logs = []
        
        for line in lines:
            # 检查是否是主机监控数据
            if "主机" in line and "状态" in line and "CPU" in line:
                host_logs.append(line.strip())
            else:
                system_logs.append(line.strip())
        
        # 优先返回主机监控数据
        if host_logs:
            return jsonify({
                'logs': host_logs,
                'count': len(host_logs),
                'type': 'host_metrics'
            })
        else:
            return jsonify({
                'logs': system_logs,
                'count': len(system_logs),
                'type': 'system'
            })
    except Exception as e:
        return jsonify({
            'logs': [f"获取日志失败: {str(e)}\n"],
            'count': 1,
            'type': 'error'
        })

    """获取最近的日志，优先返回主机监控指标"""
    try:
        log_file = os.environ.get('MONITOR_LOG_FILE', 'monitor.log')
        lines = []
        
        # 确保日志文件存在
        if not os.path.exists(log_file):
            # 创建日志文件并写入初始内容
            with open(log_file, 'w') as f:
                f.write(f"{datetime.now()}: 监控日志初始化\n")
        
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            # 返回最后50行（增加行数以显示更多主机监控数据）
            lines = lines[-50:]
        
        # 如果日志为空，返回默认消息
        if not lines:
            lines = ["暂无日志记录\n"]
        
        # 解析主机监控数据
        host_logs = []
        system_logs = []
        
        for line in lines:
            # 检查是否是主机监控数据
            if "主机" in line and "状态" in line and "CPU" in line:
                host_logs.append(line.strip())
            else:
                system_logs.append(line.strip())
        
        # 优先返回主机监控数据
        if host_logs:
            return jsonify({
                'logs': host_logs,
                'count': len(host_logs),
                'type': 'host_metrics'
            })
        else:
            return jsonify({
                'logs': system_logs,
                'count': len(system_logs),
                'type': 'system'
            })
    except Exception as e:
        return jsonify({
            'logs': [f"获取日志失败: {str(e)}\n"],
            'count': 1,
            'type': 'error'
        })

@app.route('/get_parsed_logs')
def get_parsed_logs():
    """获取解析后的日志数据，提供更结构化的格式"""
    try:
        log_file = os.environ.get('MONITOR_LOG_FILE', 'monitor.log')
        lines = []
        
        # 确保日志文件存在
        if not os.path.exists(log_file):
            return jsonify({
                'logs': [],
                'count': 0,
                'type': 'empty',
                'message': '日志文件不存在'
            })
        
        with open(log_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            # 获取最后100行
            lines = lines[-100:]
        
        if not lines:
            return jsonify({
                'logs': [],
                'count': 0,
                'type': 'empty',
                'message': '日志文件为空'
            })
        
        # 解析不同类型的日志
        parsed_logs = []
        host_data = {}  # 按主机分组
        current_host = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # 解析主机监控数据（带框框的格式）
            if "主机:" in line:
                ip_match = re.search(r'主机:\s*([\d.]+)', line)
                if ip_match:
                    ip = ip_match.group(1)
                    if ip not in host_data:
                        host_data[ip] = {
                            'ip': ip,
                            'entries': []
                        }
                    current_host = ip
                    continue
            
            # 解析状态行
            if "状态:" in line and ("✅" in line or "❌" in line):
                status_match = re.search(r'状态:\s*(\S+)', line)
                if status_match:
                    status = status_match.group(1)
                    if current_host in host_data:
                        host_data[current_host]['status'] = '在线' if '✅' in status else '离线'
                continue
                
            # 解析指标行
            if "CPU:" in line or "内存:" in line or "磁盘:" in line:
                cpu_match = re.search(r'CPU:\s*([\d.]+)%', line)
                mem_match = re.search(r'内存:\s*([\d.]+)%', line)
                disk_match = re.search(r'磁盘:\s*([\d.]+)%', line)
                
                entry = {}
                if cpu_match:
                    entry['cpu'] = float(cpu_match.group(1))
                if mem_match:
                    entry['memory'] = float(mem_match.group(1))
                if disk_match:
                    entry['disk'] = float(disk_match.group(1))
                    
                if entry and current_host in host_data:
                    # 提取时间戳（如果有）
                    time_match = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line)
                    if time_match:
                        entry['timestamp'] = time_match.group(1)
                    else:
                        entry['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                    host_data[current_host]['entries'].append(entry)
                continue
                
            # 解析网络流量行
            if "网络RX:" in line or "网络TX:" in line:
                rx_match = re.search(r'网络RX:\s*([\d.]+)(\w+)', line)
                tx_match = re.search(r'网络TX:\s*([\d.]+)(\w+)', line)
                
                if rx_match or tx_match:
                    entry = {}
                    if rx_match:
                        rx_value = float(rx_match.group(1))
                        rx_unit = rx_match.group(2)
                        # 统一转换为MB
                        if rx_unit == 'KB':
                            entry['network_rx'] = rx_value / 1024
                        elif rx_unit == 'GB':
                            entry['network_rx'] = rx_value * 1024
                        else:  # MB or unknown
                            entry['network_rx'] = rx_value
                            
                    if tx_match:
                        tx_value = float(tx_match.group(1))
                        tx_unit = tx_match.group(2)
                        # 统一转换为MB
                        if tx_unit == 'KB':
                            entry['network_tx'] = tx_value / 1024
                        elif tx_unit == 'GB':
                            entry['network_tx'] = tx_value * 1024
                        else:  # MB or unknown
                            entry['network_tx'] = tx_value
                            
                    if entry and current_host in host_data and host_data[current_host]['entries']:
                        # 将网络数据添加到最新的一条记录中
                        latest_entry = host_data[current_host]['entries'][-1]
                        latest_entry.update(entry)
                continue
                
            # 解析系统日志
            if "监控数据采集完成" in line or "ERROR" in line or "INFO" in line:
                time_match = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line)
                timestamp = time_match.group(1) if time_match else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                parsed_logs.append({
                    'type': 'system',
                    'timestamp': timestamp,
                    'message': line,
                    'level': 'error' if 'ERROR' in line else 'info'
                })
        
        # 转换主机数据为列表
        for ip, data in host_data.items():
            if data['entries']:
                # 获取最新的条目
                latest_entry = data['entries'][-1]
                parsed_logs.append({
                    'type': 'host_metrics',
                    'ip': data['ip'],
                    'status': data.get('status', '未知'),
                    'timestamp': latest_entry.get('timestamp'),
                    'cpu': latest_entry.get('cpu', 0),
                    'memory': latest_entry.get('memory', 0),
                    'disk': latest_entry.get('disk', 0),
                    'network_rx': latest_entry.get('network_rx', 0),
                    'network_tx': latest_entry.get('network_tx', 0)
                })
        
        # 按时间戳排序
        parsed_logs.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
        
        return jsonify({
            'logs': parsed_logs[:50],  # 返回最新50条
            'count': len(parsed_logs),
            'type': 'parsed'
        })
        
    except Exception as e:
        return jsonify({
            'logs': [],
            'count': 0,
            'type': 'error',
            'message': f"解析日志失败: {str(e)}"
        })


@app.route('/api/host_monitor_configs')
def api_host_monitor_configs():
    """获取主机监控配置"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 创建主机监控配置表（如果不存在）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS host_monitor_configs (
                ip TEXT NOT NULL,
                user TEXT NOT NULL,
                port INTEGER NOT NULL,
                monitor_cpu INTEGER DEFAULT 1,
                monitor_memory INTEGER DEFAULT 1,
                monitor_disk INTEGER DEFAULT 1,
                monitor_network INTEGER DEFAULT 1,
                status TEXT DEFAULT 'active',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (ip, user, port)
            )
        ''')
        
        # 获取所有配置
        cursor.execute('''
            SELECT h.ip, h.user, h.port, 
                   COALESCE(c.monitor_cpu, 1) as monitor_cpu,
                   COALESCE(c.monitor_memory, 1) as monitor_memory,
                   COALESCE(c.monitor_disk, 1) as monitor_disk,
                   COALESCE(c.monitor_network, 1) as monitor_network,
                   COALESCE(c.status, 'active') as status
            FROM hosts h
            LEFT JOIN host_monitor_configs c ON h.ip = c.ip AND h.user = c.user AND h.port = c.port
        ''')
        
        configs = cursor.fetchall()
        conn.close()
        
        # 转换为前端需要的格式
        result = {}
        for config in configs:
            ip, user, port, cpu, memory, disk, network, status = config
            key = f"{ip}-{user}-{port}"
            result[key] = {
                'monitor_cpu': bool(cpu),
                'monitor_memory': bool(memory),
                'monitor_disk': bool(disk),
                'monitor_network': bool(network),
                'status': status
            }
        
        return jsonify({'success': True, 'configs': result})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/get_current_config', methods=['GET'])
def get_current_config():
    """获取当前监控配置"""
    try:
        # 从配置文件或数据库读取当前配置
        config = {
            'interval': get_monitor_interval(),
            'cleanup_days': get_cleanup_days(),
            'monitor_cpu': True,
            'monitor_memory': True,
            'monitor_disk': True,
            'monitor_network': True
        }
        return jsonify({'success': True, **config})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/update_global_config', methods=['POST'])
def update_global_config():
    """更新全局监控配置"""
    try:
        data = request.json
        interval = data.get('interval', 300)
        monitor_cpu = data.get('cpu', True)
        monitor_memory = data.get('memory', True)
        monitor_disk = data.get('disk', True)
        monitor_network = data.get('network', True)
        
        # 更新环境变量
        os.environ['MONITOR_INTERVAL'] = str(interval)
        os.environ['MONITOR_CPU'] = str(monitor_cpu)
        os.environ['MONITOR_MEMORY'] = str(monitor_memory)
        os.environ['MONITOR_DISK'] = str(monitor_disk)
        os.environ['MONITOR_NETWORK'] = str(monitor_network)
        
        return jsonify({
            'success': True,
            'message': '全局配置更新成功'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/trigger_collection', methods=['POST'])
def trigger_collection():
    """手动触发数据采集"""
    try:
        # 直接调用采集函数
        collect_monitoring_data()
        
        return jsonify({
            'success': True,
            'message': '采集任务已完成'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/clear_logs', methods=['POST'])
def clear_logs():
    """清空日志文件"""
    try:
        log_file = os.environ.get('MONITOR_LOG_FILE', 'monitor.log')
        if os.path.exists(log_file):
            with open(log_file, 'w') as f:
                f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S,%f')} - monitor_scheduler - INFO - 日志已清空\n")
        
        return jsonify({
            'success': True,
            'message': '日志已清空'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })

@app.route('/save_config', methods=['POST'])
def save_config():
    """保存监控配置"""
    try:
        data = request.get_json()
        
        # 更新环境变量
        os.environ['MONITOR_INTERVAL'] = str(data.get('interval', 300))
        os.environ['MONITOR_CLEANUP_DAYS'] = str(data.get('cleanup_days', 30))
        os.environ['MONITOR_AUTO_CLEANUP'] = 'True' if data.get('auto_cleanup', True) else 'False'
        
        # 重启监控线程以应用新配置
        global monitor_thread
        if 'monitor_thread' in globals() and monitor_thread.is_alive():
            # 停止当前监控线程（通过设置标志位）
            if 'stop_monitoring' in globals():
                stop_monitoring = True
            else:
                globals()['stop_monitoring'] = True
            
            # 等待线程结束
            monitor_thread.join(timeout=5)
            
            # 创建新的监控线程
            import importlib
            import monitor_config
            importlib.reload(monitor_config)
            
            globals()['stop_monitoring'] = False
            monitor_thread = threading.Thread(target=monitoring_loop, daemon=True)
            monitor_thread.start()
        
        return jsonify({
            'success': True,
            'message': '配置保存成功'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })

@app.route('/api/network_trends')
def api_network_trends():
    """获取网络流量趋势数据"""
    try:
        # 获取查询参数
        ip = request.args.get('ip', '')
        limit = int(request.args.get('limit', 10))
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 构建查询条件
        where_clause = "WHERE 1=1"
        params = []
        
        if ip:
            where_clause += " AND ip = ?"
            params.append(ip)
        
        # 获取指定主机的网络流量历史数据
        query = f"""
            SELECT ip, network_rx, network_tx, check_time
            FROM monitoring_data 
            {where_clause}
            ORDER BY check_time DESC
            LIMIT ?
        """
        
        cursor.execute(query, params + [limit])
        data = cursor.fetchall()
        
        # 反转数据，使时间顺序正确
        data.reverse()
        
        # 处理数据格式
        result = {
            'labels': [],  # 时间标签
            'datasets': [
                {
                    'label': '接收流量 (MB)',
                    'data': [],
                    'borderColor': '#4CAF50',
                    'backgroundColor': 'rgba(76, 175, 80, 0.1)',
                    'tension': 0.4
                },
                {
                    'label': '发送流量 (MB)',
                    'data': [],
                    'borderColor': '#2196F3',
                    'backgroundColor': 'rgba(33, 150, 243, 0.1)',
                    'tension': 0.4
                }
            ]
        }
        
        # 提取数据
        for row in data:
            ip, rx, tx, check_time = row
            
            # 将UTC时间转换为本地时间
            try:
                # 解析UTC时间字符串
                utc_time = datetime.strptime(check_time, '%Y-%m-%d %H:%M:%S')
                # 添加时区信息并转换为中国时区
                utc_time = utc_time.replace(tzinfo=timezone.utc)
                local_time = utc_time.astimezone(pytz.timezone('Asia/Shanghai'))
                # 格式化显示时间
                time_str = local_time.strftime('%H:%M')
            except Exception as e:
                # 如果转换失败，使用原来的简单截取方法
                print(f"时间转换失败: {e}, 使用原始时间: {check_time}")
                time_str = check_time[11:16]  # 回退到原来的处理方式
            
            result['labels'].append(time_str)
            result['datasets'][0]['data'].append(round(rx, 2))
            result['datasets'][1]['data'].append(round(tx, 2))

        
        # 获取所有主机IP用于下拉选择
        cursor.execute("SELECT DISTINCT ip FROM monitoring_data WHERE network_rx > 0 OR network_tx > 0")
        available_ips = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': result,
            'available_ips': available_ips,
            'current_ip': ip if ip else '所有主机'
        })
        
    except Exception as e:
        print(f"获取网络流量趋势失败: {e}")
        return jsonify({
            'success': False,
            'message': str(e),
            'data': {
                'labels': [],
                'datasets': [
                    {'label': '接收流量', 'data': []},
                    {'label': '发送流量', 'data': []}
                ]
            },
            'available_ips': []
        })


@app.route('/alerts')
def alerts_page():
    """告警管理页面"""
    return render_template('alerts.html')

@app.route('/historical_data')
def historical_data_page():
    """历史数据查询页面"""
    return render_template('historical_data.html')

# ========== 告警管理API ==========
@app.route('/api/alert_rules', methods=['GET'])
def api_get_alert_rules():
    """获取所有告警规则"""
    rules = get_alert_rules()
    rules_list = []
    for rule in rules:
        rules_list.append({
            'id': rule[0],
            'rule_name': rule[1],
            'host_ip': rule[2],
            'metric_type': rule[3],
            'threshold_value': rule[4],
            'comparison_operator': rule[5],
            'is_active': bool(rule[6]),
            'created_at': rule[7]
        })
    return jsonify(rules_list)

@app.route('/api/alert_rules', methods=['POST'])
def api_create_alert_rule():
    """创建告警规则"""
    data = request.get_json()
    rule_name = data.get('rule_name')
    host_ip = data.get('host_ip')
    metric_type = data.get('metric_type')
    threshold_value = data.get('threshold_value')
    comparison_operator = data.get('comparison_operator', '>')
    
    if not all([rule_name, host_ip, metric_type, threshold_value]):
        return jsonify({'success': False, 'message': '缺少必要参数'})
    
    success = create_alert_rule(rule_name, host_ip, metric_type, float(threshold_value), comparison_operator)
    if success:
        return jsonify({'success': True, 'message': '告警规则创建成功'})
    else:
        return jsonify({'success': False, 'message': '创建告警规则失败'})

@app.route('/api/alert_rules/<int:rule_id>', methods=['DELETE'])
def api_delete_alert_rule(rule_id):
    """删除告警规则"""
    success = delete_alert_rule(rule_id)
    if success:
        return jsonify({'success': True, 'message': '告警规则删除成功'})
    else:
        return jsonify({'success': False, 'message': '删除告警规则失败'})

@app.route('/api/alerts/active', methods=['GET'])
def api_get_active_alerts():
    """获取未解决的活跃告警"""
    alerts = get_active_alerts()
    alerts_list = []
    for alert in alerts:
        alerts_list.append({
            'id': alert[0],
            'rule_id': alert[1],
            'host_ip': alert[2],
            'metric_type': alert[3],
            'current_value': alert[4],
            'threshold_value': alert[5],
            'message': alert[6],
            'severity': alert[7],
            'created_at': alert[8]
        })
    return jsonify(alerts_list)

@app.route('/api/alerts/<int:alert_id>/resolve', methods=['POST'])
def api_resolve_alert(alert_id):
    """解决告警"""
    success = resolve_alert(alert_id)
    if success:
        return jsonify({'success': True, 'message': '告警已解决'})
    else:
        return jsonify({'success': False, 'message': '解决告警失败'})

# ========== 修复后的历史数据查询API ==========
@app.route('/api/historical_data', methods=['GET'])
def api_get_historical_data():
    """获取历史监控数据 - 修复版本"""
    try:
        # 获取查询参数
        ip = request.args.get('ip')
        start_time = request.args.get('start_time')
        end_time = request.args.get('end_time')
        metric_type = request.args.get('metric_type', 'all')  # cpu, memory, disk, network, or all
        
        # 参数验证
        if not ip or not start_time or not end_time:
            return jsonify({'success': False, 'message': '缺少必要参数: IP地址、开始时间和结束时间'})
        
        # 验证IP地址格式
        ip_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        if not ip_pattern.match(ip):
            return jsonify({'success': False, 'message': 'IP地址格式不正确'})
        
        # 验证metric_type参数
        valid_metrics = ['all', 'cpu', 'memory', 'disk', 'network']
        if metric_type not in valid_metrics:
            return jsonify({'success': False, 'message': 'metric_type参数无效，可选值: all, cpu, memory, disk, network'})
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 安全构建查询 - 避免SQL注入
        if metric_type == 'all':
            # 查询所有指标
            cursor.execute('''
                SELECT cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online, check_time
                FROM monitoring_data
                WHERE ip = ? AND check_time BETWEEN ? AND ?
                ORDER BY check_time ASC
            ''', (ip, start_time, end_time))
        else:
            # 根据metric_type安全选择字段
            metric_field = {
                'cpu': 'cpu_usage',
                'memory': 'memory_usage', 
                'disk': 'disk_usage',
                'network': 'network_rx, network_tx'
            }.get(metric_type)
            
            if metric_type == 'network':
                cursor.execute(f'''
                    SELECT {metric_field}, check_time
                    FROM monitoring_data
                    WHERE ip = ? AND check_time BETWEEN ? AND ?
                    ORDER BY check_time ASC
                ''', (ip, start_time, end_time))
            else:
                cursor.execute(f'''
                    SELECT {metric_field}, check_time
                    FROM monitoring_data
                    WHERE ip = ? AND check_time BETWEEN ? AND ?
                    ORDER BY check_time ASC
                ''', (ip, start_time, end_time))
        
        data = cursor.fetchall()
        conn.close()
        
        # 格式化返回数据为前端友好的格式
        formatted_data = []
        if metric_type == 'all':
            for row in data:
                cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online, check_time = row
                formatted_data.append({
                    'check_time': check_time,
                    'cpu_usage': round(cpu_usage, 2) if cpu_usage is not None else 0,
                    'memory_usage': round(memory_usage, 2) if memory_usage is not None else 0,
                    'disk_usage': round(disk_usage, 2) if disk_usage is not None else 0,
                    'network_rx': round(network_rx, 2) if network_rx is not None else 0,
                    'network_tx': round(network_tx, 2) if network_tx is not None else 0,
                    'is_online': bool(is_online) if is_online is not None else False
                })
        elif metric_type == 'network':
            for row in data:
                network_rx, network_tx, check_time = row
                formatted_data.append({
                    'check_time': check_time,
                    'network_rx': round(network_rx, 2) if network_rx is not None else 0,
                    'network_tx': round(network_tx, 2) if network_tx is not None else 0
                })
        else:
            for row in data:
                value, check_time = row
                formatted_data.append({
                    'check_time': check_time,
                    metric_type: round(value, 2) if value is not None else 0
                })
        
        return jsonify({
            'success': True, 
            'data': formatted_data,
            'total_count': len(formatted_data)
        })
        
    except sqlite3.Error as e:
        print(f"查询历史数据数据库错误: {e}")
        return jsonify({'success': False, 'message': f'数据库查询失败: {str(e)}'})
    except Exception as e:
        print(f"查询历史数据未知错误: {e}")
        return jsonify({'success': False, 'message': f'查询失败: {str(e)}'})



@app.route('/api/network_summary')
def api_network_summary():
    """获取网络流量汇总数据"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # 获取最近24小时的网络流量汇总
        cursor.execute('''
            SELECT
                ip,
                AVG(network_rx) as avg_rx,
                AVG(network_tx) as avg_tx,
                MAX(network_rx) as max_rx,
                MAX(network_tx) as max_tx,
                COUNT(*) as data_points
            FROM monitoring_data
            WHERE check_time >= datetime('now', '-1 day')
            GROUP BY ip
        ''')

        data = cursor.fetchall()
        conn.close()

        result = {
            'success': True,
            'summary': []
        }

        for row in data:
            ip, avg_rx, avg_tx, max_rx, max_tx, data_points = row
            result['summary'].append({
                'ip': ip,
                'avg_rx': round(avg_rx, 2) if avg_rx is not None else 0,
                'avg_tx': round(avg_tx, 2) if avg_tx is not None else 0,
                'max_rx': round(max_rx, 2) if max_rx is not None else 0,
                'max_tx': round(max_tx, 2) if max_tx is not None else 0,
                'data_points': data_points
            })

        return jsonify(result)

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })

# 添加邮件通知功能（可选）
def send_alert_notification(alert_data):
    """发送告警通知（邮件/钉钉/企业微信）"""
    # 这里可以实现邮件、钉钉、企业微信等通知方式
    # 示例：邮件通知
    try:
        # 需要配置SMTP服务器信息
        # import smtplib
        # from email.mime.text import MIMEText
        # 实现邮件发送逻辑
        print(f"发送告警通知: {alert_data['message']}")
        return True
    except Exception as e:
        print(f"发送通知失败: {e}")
        return False



@app.route('/api/export/excel')
def api_export_excel():
    """导出监控数据为Excel格式"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        ip_filter = request.args.get('ip', '')
        
        # 参数验证 - 修复逻辑
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': '请选择开始日期和结束日期'}), 400
        
        # 验证日期格式
        try:
            from datetime import datetime
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': '日期格式错误，请使用YYYY-MM-DD格式'}), 400
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 构建查询条件 - 修复WHERE子句逻辑
        conditions = []
        params = []
        
        conditions.append("DATE(check_time) BETWEEN ? AND ?")
        params.extend([start_date, end_date])
        
        if ip_filter:
            conditions.append("ip = ?")
            params.append(ip_filter)
            
        where_clause = "WHERE " + " AND ".join(conditions)
        
        # 查询数据
        query = f"""
            SELECT ip, cpu_usage, memory_usage, disk_usage, 
                network_rx, network_tx, is_online, check_time
            FROM monitoring_data 
            {where_clause}
            ORDER BY check_time DESC
            LIMIT 1000
        """
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()

        if not data:
            return jsonify({'success': False, 'message': '没有找到符合条件的监控数据'}), 404
        
        # 检查openpyxl是否安装
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime
            from io import BytesIO
            from urllib.parse import quote
        except ImportError as e:
            return jsonify({
                'success': False, 
                'message': f'Excel导出功能需要安装openpyxl库: {str(e)}'
            }), 500
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "监控数据"
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置表头
        headers = ['主机IP', 'CPU使用率(%)', '内存使用率(%)', '磁盘使用率(%)',
                  '网络接收(MB)', '网络发送(MB)', '在线状态', '记录时间']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            ip, cpu_usage, memory_usage, disk_usage, network_rx, network_tx, is_online, check_time = row_data
            
            # 处理NULL值
            cpu_usage = cpu_usage if cpu_usage is not None else 0
            memory_usage = memory_usage if memory_usage is not None else 0
            disk_usage = disk_usage if disk_usage is not None else 0
            network_rx = network_rx if network_rx is not None else 0
            network_tx = network_tx if network_tx is not None else 0
            status = '在线' if is_online else '离线'
            
            ws.cell(row=row_idx, column=1, value=ip)
            ws.cell(row=row_idx, column=2, value=cpu_usage)
            ws.cell(row=row_idx, column=3, value=memory_usage)
            ws.cell(row=row_idx, column=4, value=disk_usage)
            ws.cell(row=row_idx, column=5, value=network_rx)
            ws.cell(row=row_idx, column=6, value=network_tx)
            ws.cell(row=row_idx, column=7, value=status)
            ws.cell(row=row_idx, column=8, value=check_time)
        
        # 调整列宽
        column_widths = [15, 12, 12, 12, 12, 12, 10, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名 - 修复文件名编码
        filename = f"监控数据_{start_date}_至_{end_date}.xlsx"
        if ip_filter:
            filename = f"监控数据_{ip_filter}_{start_date}_至_{end_date}.xlsx"
        
        encoded_filename = quote(filename.encode('utf-8'))

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'

        # 添加调试日志
        print(f"Excel导出成功：找到 {len(data)} 条记录，文件大小: {len(output.getvalue())} 字节")

        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel导出错误: {error_details}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

# 报告生成API
@app.route('/api/export/report')
def api_export_report():
    """统一监控报告生成API"""
    try:
        from datetime import datetime
        from io import BytesIO
        
        # 获取并验证请求参数
        report_type = request.args.get('type', 'html').lower()
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        host_ip = request.args.get('host_ip', '')
        template = request.args.get('template', 'summary')
        title = request.args.get('title', '服务器监控报告')
        
        # 参数验证
        if not start_date or not end_date:
            return jsonify({
                'success': False, 
                'message': '请提供开始日期和结束日期',
                'required_params': ['start_date', 'end_date']
            })
        
        # 验证日期格式
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({
                'success': False, 
                'message': '日期格式错误，请使用YYYY-MM-DD格式',
                'example': '2024-01-01'
            })
        
        # 验证报告类型
        valid_report_types = ['html']
        if report_type not in valid_report_types:
            return jsonify({
                'success': False, 
                'message': f'不支持的报告格式: {report_type}',
                'supported_types': valid_report_types
            })
        
        # 验证模板类型
        valid_templates = ['summary', 'detailed', 'performance', 'trends']
        if template not in valid_templates:
            return jsonify({
                'success': False, 
                'message': f'不支持的模板类型: {template}',
                'supported_templates': valid_templates
            })
        
        # 根据报告类型调用相应的生成函数
        if report_type == 'html':
            result = generate_html_report_v2(start_date, end_date, host_ip, template, title)
        else:
            return jsonify({'success': False, 'message': '未知的报告类型'})
        
        return result
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"报告生成错误: {error_details}")
        
        return jsonify({
            'success': False, 
            'message': f'报告生成失败: {str(e)}',
            'error_details': str(e)
        })

def generate_html_report_v2(start_date, end_date, host_ip='', template='summary', title='服务器监控报告'):
    """增强版HTML报告生成函数"""
    try:
        # 构建查询条件
        conditions = []
        params = []
        
        conditions.append("DATE(check_time) BETWEEN ? AND ?")
        params.extend([start_date, end_date])
        
        if host_ip:
            conditions.append("ip = ?")
            params.append(host_ip)
        
        where_clause = "WHERE " + " AND ".join(conditions)
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 获取基础统计信息
        cursor.execute(f"""
            SELECT 
                COUNT(DISTINCT ip) as host_count,
                AVG(cpu_usage) as avg_cpu,
                AVG(memory_usage) as avg_memory,
                AVG(disk_usage) as avg_disk,
                AVG(network_rx) as avg_rx,
                AVG(network_tx) as avg_tx,
                COUNT(*) as total_records,
                COUNT(DISTINCT CASE WHEN is_online = 1 THEN ip END) as online_hosts,
                MAX(cpu_usage) as max_cpu,
                MAX(memory_usage) as max_memory,
                MAX(disk_usage) as max_disk
            FROM monitoring_data 
            {where_clause}
        """, params)
        
        stats = cursor.fetchone()
        
        # 获取详细数据用于图表
        cursor.execute(f"""
            SELECT ip, cpu_usage, memory_usage, disk_usage, network_rx, network_tx, 
                   is_online, check_time
            FROM monitoring_data 
            {where_clause}
            ORDER BY check_time DESC
            LIMIT 500
        """, params)
        
        detailed_data = cursor.fetchall()
        conn.close()
        
        # 生成HTML报告
        return generate_enhanced_html_report(stats, detailed_data, start_date, end_date, host_ip, template, title)
        
    except Exception as e:
        raise Exception(f"HTML报告生成失败: {str(e)}")

def generate_enhanced_html_report(stats, detailed_data, start_date, end_date, host_ip, template, title):
    """生成增强的HTML报告"""
    from datetime import datetime
    from io import BytesIO
    
    # 解析统计信息
    host_count, avg_cpu, avg_memory, avg_disk, avg_rx, avg_tx, total_records, online_hosts, max_cpu, max_memory, max_disk = stats
    
    # 生成HTML内容
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{title}</title>
        <meta charset="utf-8">
        <style>
            body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; line-height: 1.6; }}
            .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
            .section {{ margin: 25px 0; padding: 20px; background: white; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            .stat-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 20px; margin: 20px 0; }}
            .stat-item {{ padding: 20px; background: #f8f9fa; border-radius: 8px; text-align: center; border-left: 4px solid #007bff; }}
            .stat-value {{ font-size: 28px; font-weight: bold; color: #007bff; margin-bottom: 5px; }}
            .stat-label {{ font-size: 14px; color: #6c757d; }}
            .recommendation {{ margin: 12px 0; padding: 15px; background: #d4edda; border-left: 4px solid #28a745; border-radius: 4px; }}
            .data-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            .data-table th, .data-table td {{ padding: 12px; text-align: left; border-bottom: 1px solid #dee2e6; }}
            .data-table th {{ background: #f8f9fa; font-weight: 600; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>📊 {title}</h1>
            <p>📅 统计期间: {start_date} 至 {end_date}</p>
            <p>🖥️ 监控范围: {f"主机 {host_ip}" if host_ip else "所有主机"} | 总数: {host_count}台</p>
            <p>⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="section">
            <h2>📈 性能概览</h2>
            <div class="stat-grid">
                <div class="stat-item">
                    <div class="stat-value">{avg_cpu:.1f}%</div>
                    <div class="stat-label">平均CPU使用率</div>
                    <div style="font-size: 12px; color: #888;">峰值: {max_cpu:.1f}%</div>
                </div>
                <div class="stat-item">
                    <div class="stat-value">{avg_memory:.1f}%</div>
                    <div class="stat-label">平均内存使用率</div>
                    <div style="font-size: 12px; color: #888;">峰值: {max_memory:.1f}%</div>
                </div>
                <div class="stat-item">
                    <div class="stat-value">{avg_disk:.1f}%</div>
                    <div class="stat-label">平均磁盘使用率</div>
                    <div style="font-size: 12px; color: #888;">峰值: {max_disk:.1f}%</div>
                </div>
                <div class="stat-item">
                    <div class="stat-value">{online_hosts}/{host_count}</div>
                    <div class="stat-label">在线主机</div>
                    <div style="font-size: 12px; color: #888;">可用性: {(online_hosts/host_count*100 if host_count>0 else 0):.1f}%</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2>📋 详细数据</h2>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>主机IP</th>
                        <th>CPU%</th>
                        <th>内存%</th>
                        <th>磁盘%</th>
                        <th>网络RX</th>
                        <th>网络TX</th>
                        <th>状态</th>
                        <th>检查时间</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(
                        f'<tr><td>{row[0]}</td><td>{row[1]:.1f}</td><td>{row[2]:.1f}</td>'
                        f'<td>{row[3]:.1f}</td><td>{row[4]:.1f}MB</td><td>{row[5]:.1f}MB</td>'
                        f'<td>{"✅在线" if row[6] else "❌离线"}</td><td>{row[7]}</td></tr>'
                        for row in detailed_data[:50]
                    )}
                </tbody>
            </table>
            <p style="text-align: center; color: #6c757d; margin-top: 10px;">
                显示前50条记录，共{len(detailed_data)}条记录
            </p>
        </div>
    </body>
    </html>
    """
    
    # 返回HTML文件
    output = BytesIO(html_content.encode('utf-8'))
    
    return send_file(
        output,
        download_name=f"{title}_{start_date}_to_{end_date}.html",
        as_attachment=True,
        mimetype='text/html'
    )

@app.route('/api/export/csv')
def api_export_csv():
    """导出监控数据为CSV格式"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        ip_filter = request.args.get('ip', '')
        
        # 参数验证
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': '请选择开始日期和结束日期'})
        
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 构建查询条件
        conditions = []
        params = []
        
        if start_date:
            conditions.append("DATE(check_time) >= ?")
            params.append(start_date)
            
        if end_date:
            conditions.append("DATE(check_time) <= ?")
            params.append(end_date)
            
        if ip_filter:
            conditions.append("ip = ?")
            params.append(ip_filter)
            
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        
        # 查询数据
        query = f"""
            SELECT ip, cpu_usage, memory_usage, disk_usage, 
                network_rx, network_tx, is_online, check_time
            FROM monitoring_data 
            {where_clause}
            ORDER BY check_time DESC
            LIMIT 1000
        """
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return jsonify({'success': False, 'message': '没有找到符合条件的监控数据'})
        
        # 生成CSV内容
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['主机IP', 'CPU使用率%', '内存使用率%', '磁盘使用率%', 
                        '接收流量(MB)', '发送流量(MB)', '在线状态', '检查时间'])
        
        # 写入数据
        for row in data:
            writer.writerow(row)
        
        # 返回CSV文件
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=monitor_data_{start_date}_to_{end_date}.csv'
        
        return response
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

if __name__ == '__main__':
    # 初始化数据库
    init_db()
    
    # 启动监控线程
    monitor_thread = threading.Thread(target=monitoring_loop, daemon=True)
    monitor_thread.start()
    
    app.run(host='0.0.0.0', port=80)